"""Extrae de las pautas de mantención los repuestos que deben quedar marcados InStock.

Genera `apps/api/src/data/pautas_instock.csv`, la lista maestra de part numbers que
la plataforma tiene que mantener siempre en stock en las sucursales con taller
(ver `services/instock_service.py`).

Alcance (decidido con Abastecimiento, jul-2026):
  - FORD: Transit, Ranger y F-150, años modelo **2023 y 2024**. Es el esquema
    vigente de la pauta (hojas "IOLM 16K 23/24" y "20K"); las hojas "10K" son el
    plan anterior (2021-2022) y quedan fuera. La pauta no trae año modelo 2025 ni
    2026: 2024 es el ultimo que publica el fabricante.
  - HYUNDAI: Accent, Grand i10 e i20, **todas sus generaciones**. Acá no se filtra
    por año: el libro organiza las pautas por generación (RB, HCi, BN7i, BA, Ai3,
    IB, Bi3…) y el criterio es el modelo, no el año de fabricación.

No entran lubricantes ni consumibles (aceite, fluido de frenos, refrigerante,
limpiaparabrisas, "compra en plaza"): se venden por litro y un mínimo de "2
unidades" no significa nada sobre ellos.

Uso:
    python scripts/extraer_instock_pautas.py [--pautas DIR] [--catalogo CSV]

El CSV de catálogo es opcional y solo sirve para el reporte de cobertura (cuántos
part numbers existen hoy como código de producto). El cruce real lo hace
`src.jobs.cargar_instock` contra la base.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[1]
PAUTAS_DEFECTO = Path(
    r"C:\Users\icalderon\OneDrive - Curifor S.A\Documentos\Desarrollos"
    r"\Bases de datos\Pautas de mantención"
)
SALIDA = RAIZ / "apps" / "api" / "src" / "data" / "pautas_instock.csv"
CATALOGO_DEFECTO = RAIZ / "lista productos.csv"

FORD_XLSM = "Pauta Servicio Ford - 17-06-2026.xlsm"
HYUNDAI_XLSX = "Pauta Mantención Hyundai 2026.xlsx"

# --- Ford ------------------------------------------------------------------
# Hojas del esquema vigente. Las "10K" son el plan anterior (2021-2022).
FORD_HOJAS = (
    "Plan Mantenimiento IOLM 16K 23",
    "Plan Mantenimiento IOLM 16K 24",
    "Plan Mantenimiento 20K",
)
# Años modelo pedidos. Va aparte de la seleccion de hojas a proposito: la hoja
# "20K" mezcla bloques de 2023 y 2024, y si el fabricante agrega un año nuevo al
# libro no se cuela solo porque comparte hoja.
FORD_ANIOS = (2023, 2024)
# Familias pedidas. El libro separa el Transit por carrocería (V362 Custom, V363,
# V710 Custom, y el BEV bajo "Transit" pelado): todas son Transit.
FORD_FAMILIAS = ("transit", "ranger", "f-150")

# --- Hyundai ---------------------------------------------------------------
# Modelos pedidos, con TODAS sus generaciones: el criterio es el modelo, no el
# año. Se toma cualquier hoja de costos que empiece con uno de estos nombres (las
# hojas "plan" solo traen la matriz R/I/C de operaciones, sin códigos). El prefijo
# evita falsos positivos: "i30", "i30 N" e "Ioniq" no matchean "i20" ni "i10".
HYUNDAI_FAMILIAS = {
    "accent": "Accent",
    "grand i10": "Grand i10",
    "i10": "Grand i10",
    "i20": "i20",
}

# Consumibles: no son repuestos de stock unitario.
OPERACIONES_EXCLUIDAS = (
    "aceite", "fluido", "liquido", "líquido", "limpiador", "refrigerante",
    "grasa", "aditivo", "urea", "adblue",
)
# Piezas: manda sobre la lista de arriba. Sin esto, "FILTRO ACEITE MOTOR" (Hyundai)
# se descartaba por contener "aceite", que es la pieza que más importa tener.
OPERACIONES_PIEZA = (
    "filtro", "correa", "golilla", "bujia", "bujía", "reten", "retén", "kit",
    "pastilla", "disco", "escobilla", "manguera", "tapon", "tapón", "sello",
    "empaque", "junta", "balata", "amortiguador", "rodamiento", "polea",
)
# Valores que la pauta usa como marcador, no como código de repuesto.
PARTES_EXCLUIDAS = ("compra en plaza", "n/a", "na", "-", "no aplica", "s/c")


def _norm(texto: str | None) -> str:
    """Código comparable: sin espacios, guiones ni barras, en mayúsculas."""
    return re.sub(r"[^A-Z0-9]", "", (texto or "").upper())


def _sin_tildes(texto: str) -> str:
    tabla = str.maketrans("áéíóúÁÉÍÓÚñÑ", "aeiouAEIOUnN")
    return texto.translate(tabla)


def _es_consumible(operacion: str) -> bool:
    op = _sin_tildes((operacion or "").lower())
    if any(_sin_tildes(x) in op for x in OPERACIONES_PIEZA):
        return False
    return any(_sin_tildes(x) in op for x in OPERACIONES_EXCLUIDAS)


def _parte_valida(parte: str) -> bool:
    p = (parte or "").strip().lower()
    if not p or p in PARTES_EXCLUIDAS:
        return False
    # Un código de repuesto siempre tiene dígitos y nada de espacios internos.
    return bool(re.search(r"\d", p)) and " " not in p.strip()


class Registro:
    """Un part number de la pauta, con los modelos donde aparece."""

    def __init__(self, part_number: str, marca: str, operacion: str):
        self.part_number = part_number
        self.marca = marca
        self.operacion = operacion
        self.modelos: list[str] = []
        self.detalle: list[str] = []

    def agregar(self, modelo: str, detalle: str) -> None:
        if modelo not in self.modelos:
            self.modelos.append(modelo)
        if detalle not in self.detalle:
            self.detalle.append(detalle)


def _acumular(acum: dict, part_number: str, marca: str, operacion: str,
              modelo: str, detalle: str, descartados: list) -> None:
    if _es_consumible(operacion):
        descartados.append(("consumible", marca, modelo, operacion, part_number))
        return
    if not _parte_valida(part_number):
        descartados.append(("codigo no util", marca, modelo, operacion, part_number))
        return
    clave = (marca, _norm(part_number))
    reg = acum.get(clave)
    if reg is None:
        reg = Registro(part_number.strip(), marca, operacion.strip())
        acum[clave] = reg
    reg.agregar(modelo, detalle)


def extraer_ford(ruta: Path, acum: dict, descartados: list) -> None:
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        for hoja in FORD_HOJAS:
            if hoja not in wb.sheetnames:
                print(f"  AVISO: falta la hoja {hoja!r}", file=sys.stderr)
                continue
            ws = wb[hoja]
            # Estructura: cada bloque arranca con [año | modelo | versión] en las
            # columnas B..D, y debajo van las filas [operación | no parte | cant].
            anio = modelo = version = None
            for b, c, d in ws.iter_rows(min_col=2, max_col=4, values_only=True):
                if isinstance(b, (int, float)) and 2015 <= int(b) <= 2035:
                    anio, modelo = int(b), (c or "").strip() if isinstance(c, str) else ""
                    version = d.strip() if isinstance(d, str) else ""
                    continue
                if not modelo or not isinstance(b, str) or not isinstance(c, str):
                    continue
                operacion, parte = b.strip(), c.strip()
                if not operacion or operacion.lower() == "operación":
                    continue
                if anio not in FORD_ANIOS:
                    continue
                if not modelo.lower().startswith(FORD_FAMILIAS):
                    continue
                familia = "Transit" if modelo.lower().startswith("transit") else modelo
                _acumular(
                    acum, parte, "FORD", operacion, familia,
                    f"{modelo} {version} ({anio})".strip(), descartados,
                )
    finally:
        wb.close()


def _familia_hyundai(hoja: str) -> str | None:
    """Modelo pedido al que pertenece la hoja de costos, o None si no aplica."""
    nombre = _sin_tildes(hoja.strip().lower())
    if "costo" not in nombre:
        return None  # las hojas "plan" no traen part numbers
    for prefijo, familia in HYUNDAI_FAMILIAS.items():
        if nombre.startswith(prefijo):
            return familia
    return None


def extraer_hyundai(ruta: Path, acum: dict, descartados: list) -> None:
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        hojas = [(h, _familia_hyundai(h)) for h in wb.sheetnames]
        hojas = [(h, fam) for h, fam in hojas if fam]
        if not hojas:
            print("  AVISO: ninguna hoja de costos de Accent / i10 / i20", file=sys.stderr)
        for hoja, familia in hojas:
            ws = wb[hoja]
            # Estructura: la sección de repuestos arranca en la fila cuyo B dice
            # "REPUESTOS" (B=descripción, C=n° repuesto) y termina en "MATERIALES"
            # / "LUBRICANTES" / "TOTAL". El titulo de la generación esta en B1
            # ("Accent 1.5 (BN7i)") y es lo que se guarda como detalle.
            en_repuestos = False
            detalle = hoja
            n_antes = len(acum)
            for i, (b, c) in enumerate(ws.iter_rows(min_col=2, max_col=3, values_only=True), 1):
                if i == 1 and isinstance(b, str) and b.strip():
                    detalle = b.strip()
                    continue
                etiqueta = b.strip().upper() if isinstance(b, str) else ""
                if etiqueta == "REPUESTOS":
                    en_repuestos = True
                    continue
                if etiqueta in ("MATERIALES", "LUBRICANTES", "TOTAL"):
                    en_repuestos = False
                    continue
                if not en_repuestos or not isinstance(c, str):
                    continue
                _acumular(acum, c, "HYUNDAI", b if isinstance(b, str) else "",
                          familia, detalle, descartados)
            print(f"    {hoja:<34} {familia:<10} (+{len(acum) - n_antes} nuevos)")
    finally:
        wb.close()


def cobertura(acum: dict, catalogo: Path) -> dict:
    """Cuenta cuántos part numbers existen hoy como código en el catálogo del ERP."""
    if not catalogo.exists():
        return {}
    buscados = {clave[1] for clave in acum}
    encontrados: dict[str, set[str]] = {}
    with open(catalogo, encoding="latin-1", newline="") as fh:
        for fila in csv.DictReader(fh, delimiter=";"):
            producto = (fila.get("Producto") or "").strip()
            m = re.match(r"^\d+\s+(.+)$", producto)
            if not m:
                continue
            codigo = _norm(m.group(1))
            if codigo in buscados:
                encontrados.setdefault(codigo, set()).add(producto)
    return encontrados


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pautas", type=Path, default=PAUTAS_DEFECTO)
    ap.add_argument("--catalogo", type=Path, default=CATALOGO_DEFECTO)
    ap.add_argument("--salida", type=Path, default=SALIDA)
    args = ap.parse_args()

    acum: dict = {}
    descartados: list = []

    ford = args.pautas / FORD_XLSM
    hyundai = args.pautas / HYUNDAI_XLSX
    print(f"Leyendo {ford.name}…")
    extraer_ford(ford, acum, descartados)
    print(f"Leyendo {hyundai.name}…")
    extraer_hyundai(hyundai, acum, descartados)

    print(f"\nPart numbers únicos: {len(acum)}")
    for marca in ("FORD", "HYUNDAI"):
        n = sum(1 for k in acum if k[0] == marca)
        print(f"  {marca}: {n}")

    enc = cobertura(acum, args.catalogo)
    if enc:
        sin_codigo = [r for k, r in acum.items() if k[1] not in enc]
        print(f"\nCobertura contra el catálogo del ERP ({args.catalogo.name}):")
        print(f"  con código de producto: {len(acum) - len(sin_codigo)}")
        print(f"  sin código (no están en el maestro): {len(sin_codigo)}")
        for r in sin_codigo:
            print(f"    - {r.marca} {r.part_number} · {r.operacion} · {', '.join(r.modelos)}")

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    with open(args.salida, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["part_number", "marca", "modelos", "operacion", "detalle"])
        for _clave, r in sorted(acum.items(), key=lambda kv: (kv[0][0], kv[1].part_number)):
            w.writerow([
                r.part_number, r.marca, ", ".join(sorted(r.modelos)),
                r.operacion, " | ".join(r.detalle[:6]),
            ])
    print(f"\nEscrito: {args.salida}")

    if descartados:
        print(f"\nDescartados ({len(descartados)} filas de pauta):")
        vistos = set()
        for motivo, marca, modelo, op, parte in descartados:
            clave = (motivo, marca, op, parte)
            if clave in vistos:
                continue
            vistos.add(clave)
            print(f"  [{motivo}] {marca} {modelo} · {op} · {parte!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
