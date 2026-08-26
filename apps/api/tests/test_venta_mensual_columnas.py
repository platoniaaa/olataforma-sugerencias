"""Las 12 columnas de venta mensual y los promedios a 3, 6 y 12 meses.

El motor las publica con nombre POSICIONAL (`Venta Mes 01` es el ultimo mes
cerrado) y manda aparte `Periodo Ultimo Mes` para poder ponerles fecha. Los tres
puntos donde esto se rompe en silencio:

1. La cabecera del motor no calza con el alias del loader -> la columna se
   descarta con un "Columnas ignoradas (sin mapeo)" que nadie lee. Ya paso con
   `meses_con_venta_3m`.
2. La columna esta en el modelo pero no en las mini-migraciones -> en produccion
   no existe, y los servicios que se tragan la excepcion la apagan sin error.
3. El export no la tiene en LABELS -> el Excel que baja el comprador sale sin
   ella, tambien en silencio.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.pool import StaticPool

from src.db import Base, create_all
from src.models import Sugerido
from src.services import excel_loader
from src.services.excel_loader import _cast
from src.services.excel_export import etiqueta_mes, generar_excel


# --- La etiqueta del mes --------------------------------------------------------


def test_la_posicion_se_traduce_al_mes_que_corresponde():
    """01 es el ultimo mes cerrado y de ahi hacia atras, cruzando el año."""
    assert etiqueta_mes("202606", 1) == "Venta jun-26"
    assert etiqueta_mes("202606", 6) == "Venta ene-26"
    assert etiqueta_mes("202606", 7) == "Venta dic-25"
    assert etiqueta_mes("202606", 12) == "Venta jul-25"


def test_sin_periodo_no_se_inventa_un_mes():
    """El archivo se guarda y se mira semanas despues: un mes adivinado miente.

    Sin dato se devuelve None y el export cae a "Venta Mes 03", que es feo pero
    no le pone fecha a algo que no la tiene.
    """
    assert etiqueta_mes(None, 3) is None
    assert etiqueta_mes("", 3) is None
    assert etiqueta_mes("2026", 3) is None
    assert etiqueta_mes("ago-26", 3) is None


# --- Que el dato llegue de verdad -----------------------------------------------


CABECERAS = ["producto", "sucursal_id", "Venta Mes 01", "Venta Mes 12",
             "Prom Vta 3m", "Prom Vta 6m", "Prom Vta 12m", "Periodo Ultimo Mes"]


def _excel(fila: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(CABECERAS)
    ws.append(fila)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_el_loader_reconoce_las_cabeceras_del_motor(db_session):
    """Es el paso que ya fallo una vez y no avisa: cabecera sin alias = columna perdida."""
    excel_loader.cargar_sugerido(
        db_session, "sugerido.xlsx",
        _excel(["17 GK2Z9365C", "LINDEROS", 4, 2, 3.5, 1.5, 0.75, "202606"]))

    f = db_session.query(Sugerido).filter_by(producto="17 GK2Z9365C").one()
    assert f.venta_mes_01 == 4
    assert f.venta_mes_12 == 2
    assert f.prom_vta_3m == 3.5
    assert f.prom_vta_6m == 1.5
    assert f.prom_vta_12m == 0.75
    assert f.periodo_ultimo_mes == "202606"


def test_la_venta_mensual_admite_decimales(db_session):
    """Hay granel (litros, ml) y las notas de credito restan: un Integer truncaria."""
    excel_loader.cargar_sugerido(
        db_session, "sugerido.xlsx",
        _excel(["19 GRANEL", "TALCA", 2.5, -1.5, 0.5, 0.25, 0.125, "202606"]))

    f = db_session.query(Sugerido).filter_by(producto="19 GRANEL").one()
    assert f.venta_mes_01 == 2.5
    assert f.venta_mes_12 == -1.5
    assert f.prom_vta_12m == 0.125


# --- Que la columna exista en la base que ya esta creada ------------------------


@pytest.fixture()
def engine_con_tablas_viejas(monkeypatch):
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False},
                        poolclass=StaticPool)
    Base.metadata.create_all(bind=eng)
    monkeypatch.setattr("src.db.engine", eng)
    return eng


@pytest.mark.parametrize("columna", ["venta_mes_01", "venta_mes_12", "prom_vta_3m",
                                     "prom_vta_12m", "periodo_ultimo_mes"])
def test_la_migracion_repone_la_columna_en_una_tabla_que_ya_existia(
        engine_con_tablas_viejas, columna):
    """`create_all()` no agrega columnas a una tabla existente: hace falta el ALTER."""
    eng = engine_con_tablas_viejas
    with eng.begin() as cx:
        cx.execute(text(f"ALTER TABLE sugerido DROP COLUMN {columna}"))

    create_all()

    cols = {c["name"] for c in inspect(eng).get_columns("sugerido")}
    assert columna in cols, (
        f"Falta el ADD COLUMN de `{columna}` en el bloque `migraciones` de db.py: "
        "en produccion la columna no va a existir y nadie va a ver un error.")


# --- Que salga en el Excel del comprador ----------------------------------------


def test_el_export_titula_los_meses_con_su_fecha():
    """"Venta Mes 03" en un archivo guardado no le sirve a nadie tres semanas despues."""
    filas = [{"producto": "A", "venta_mes_01": 4, "venta_mes_03": 1,
              "periodo_ultimo_mes": "202606"}]

    wb = load_workbook(io.BytesIO(generar_excel(
        filas, ["producto", "venta_mes_01", "venta_mes_03"])))
    cab = [c.value for c in wb.active[1]]

    assert cab == ["Producto", "Venta jun-26", "Venta abr-26"]


def test_el_export_no_descarta_las_columnas_nuevas():
    """`generar_excel` filtra por LABELS: lo que no este ahi desaparece sin avisar."""
    columnas = ([f"venta_mes_{i:02d}" for i in range(1, 13)]
                + ["prom_vta_3m", "prom_vta_6m", "prom_vta_12m"])
    filas = [{**{c: 1 for c in columnas}, "producto": "A", "periodo_ultimo_mes": "202606"}]

    wb = load_workbook(io.BytesIO(generar_excel(filas, ["producto", *columnas])))

    assert len([c.value for c in wb.active[1]]) == len(columnas) + 1


def test_el_cast_normaliza_el_periodo():
    """El periodo tiene que quedar en "YYYYMM" venga como venga.

    Va como texto, pero segun de donde salga la celda llega entero o float, y
    `str(202607.0)` deja "202607.0": seis digitos mas basura. `etiqueta_mes` lo
    rechaza y TODA la grilla vuelve a "Venta Mes 01". No revienta nada, asi que
    nadie lo reporta como error; solo se dejan de ver los meses.
    """
    assert _cast("periodo_ultimo_mes", "202607") == "202607"
    assert _cast("periodo_ultimo_mes", 202607) == "202607"
    assert _cast("periodo_ultimo_mes", 202607.0) == "202607"
    assert _cast("periodo_ultimo_mes", " 202607 ") == "202607"


def test_un_periodo_que_no_es_un_periodo_queda_en_nulo():
    """Mejor sin etiqueta que con una fecha inventada: el Excel se guarda y se
    mira semanas despues."""
    assert _cast("periodo_ultimo_mes", "ago-26") is None
    assert _cast("periodo_ultimo_mes", "2026") is None
    assert _cast("periodo_ultimo_mes", "") is None


def test_el_periodo_sobrevive_al_viaje_por_el_archivo(db_session):
    """El camino real: el motor manda "Periodo Ultimo Mes" y tiene que llegar
    entero hasta la fila, listo para etiquetar los meses."""
    excel_loader.cargar_sugerido(
        db_session, "sugerido.xlsx",
        _excel(["17 PERIODO", "LINDEROS", 1, 1, 1, 1, 1, "202607"]))

    f = db_session.query(Sugerido).filter_by(producto="17 PERIODO").one()
    assert f.periodo_ultimo_mes == "202607"
    assert etiqueta_mes(f.periodo_ultimo_mes, 1) == "Venta jul-26"
