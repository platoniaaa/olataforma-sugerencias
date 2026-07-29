"""Las unidades salen sin decimales en el Excel.

`Total Sugerido` salia "815.138,00": el formato se elegia por el tipo Python y
esas columnas son Float en la BD aunque el valor sea entero. Ademas quedaba
incoherente con el stock por bodega (Integer), que salia "134.997" en la misma
planilla.
"""
import io

import openpyxl
from src.services import excel_export


def _hoja(rows, cols):
    return openpyxl.load_workbook(io.BytesIO(excel_export.generar_excel(rows, cols))).active


def test_las_unidades_no_llevan_decimales():
    fila = {
        "producto": "P-1",
        "total_sugerido_suc": 815138.0,     # float en la BD, entero de verdad
        "stock_activo_suc": 134997.0,
        "sugerido_compra_neto": 53.0,
        "stock_linderos": 134997,           # int: ya salia bien
        "demanda_mensual": 1265915.8333,    # este SI lleva decimales
        "costo_unitario": 3789.674,         # CLP
    }
    ws = _hoja([fila], list(fila))
    fmt = {ws.cell(row=1, column=j).value: ws.cell(row=2, column=j).number_format
           for j in range(1, ws.max_column + 1)}
    assert fmt["Total Sugerido"] == "#,##0"
    assert fmt["Stock Activo"] == "#,##0"
    assert fmt["Sugerido Compra Neto"] == "#,##0"
    assert fmt["Stock Linderos"] == "#,##0"
    # Lo que de verdad tiene decimales los conserva.
    assert fmt["Demanda Mensual"] == "#,##0.00"
    # Y el dinero sigue con formato de peso.
    assert fmt["Costo Unitario"] == '"$"#,##0'


def test_el_stock_por_bodega_y_el_stock_activo_usan_el_mismo_formato():
    fila = {"producto": "P-2", "stock_activo_suc": 10.0, "stock_talca": 10}
    ws = _hoja([fila], list(fila))
    fmt = [ws.cell(row=2, column=j).number_format for j in (2, 3)]
    assert fmt[0] == fmt[1], f"formatos distintos para lo mismo: {fmt}"
