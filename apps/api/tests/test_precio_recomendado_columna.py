"""La columna "Precio Recomendado Compra" tiene que llegar entera hasta la grilla.

El motor calcula el menor de los precios de COMPRA de FORD. Del lado de la
plataforma lo unico que hace falta es que el dato no se pierda por el camino, y
hay tres lugares donde eso pasa **sin dar ningun error**:

1. La cabecera del motor no calza con el alias de la carga -> la columna se
   descarta con un "Columnas ignoradas (sin mapeo)" que nadie lee.
2. La columna esta en el modelo pero no en las mini-migraciones -> en produccion
   no existe.
3. No esta en LABELS del export -> el Excel del comprador sale sin ella.

Ya paso con `meses_con_venta_3m`, que el motor publicaba desde siempre y la carga
tiraba a la basura.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.pool import StaticPool

from src.db import Base, create_all
from src.models import Sugerido
from src.services import excel_loader
from src.services.excel_export import CLP_COLUMNS, LABELS, generar_excel

CAMPO = "precio_recomendado_compra"


def _excel(valor) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["producto", "sucursal_id", "Precio Recomendado Compra"])
    ws.append(["17 GK2Z9365C", "LINDEROS", valor])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_la_carga_reconoce_la_cabecera_del_motor(db_session):
    excel_loader.cargar_sugerido(db_session, "sugerido.xlsx", _excel(32422))

    fila = db_session.query(Sugerido).filter_by(producto="17 GK2Z9365C").one()
    assert getattr(fila, CAMPO) == 32422


def test_un_producto_sin_precio_queda_vacio_y_no_en_cero(db_session):
    """Cero se leeria como "gratis" y ordenaria ese producto primero al ordenar
    la grilla por precio."""
    excel_loader.cargar_sugerido(db_session, "sugerido.xlsx", _excel(None))

    assert getattr(db_session.query(Sugerido).one(), CAMPO) is None


@pytest.fixture()
def engine_con_tablas_viejas(monkeypatch):
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False},
                        poolclass=StaticPool)
    Base.metadata.create_all(bind=eng)
    monkeypatch.setattr("src.db.engine", eng)
    return eng


def test_la_migracion_repone_la_columna(engine_con_tablas_viejas):
    """`create_all()` no agrega columnas a una tabla que ya existe."""
    eng = engine_con_tablas_viejas
    with eng.begin() as cx:
        cx.execute(text(f"ALTER TABLE sugerido DROP COLUMN {CAMPO}"))

    create_all()

    assert CAMPO in {c["name"] for c in inspect(eng).get_columns("sugerido")}, (
        f"Falta el ADD COLUMN de `{CAMPO}` en el bloque `migraciones` de db.py.")


def test_el_export_la_incluye_y_con_formato_de_pesos():
    """`generar_excel` filtra por LABELS: lo que no este ahi desaparece sin avisar."""
    assert CAMPO in LABELS
    assert CAMPO in CLP_COLUMNS, "sin esto sale como un numero suelto, no como precio"

    wb = load_workbook(io.BytesIO(generar_excel(
        [{"producto": "A", CAMPO: 32422}], ["producto", CAMPO])))
    hoja = wb.active

    assert [c.value for c in hoja[1]] == ["Producto", "Precio Recomendado Compra"]
    assert hoja.cell(row=2, column=2).value == 32422
    assert "$" in hoja.cell(row=2, column=2).number_format


def test_la_columna_esta_en_el_catalogo_del_frontend():
    """El frontend define las columnas en su propio archivo; si no esta ahi, el
    dato llega al navegador y no se muestra en ninguna parte."""
    from pathlib import Path

    columnas = Path(__file__).resolve().parents[2] / "web" / "lib" / "columnas.ts"
    texto = columnas.read_text(encoding="utf-8")

    assert f'key: "{CAMPO}"' in texto


# --- El tipo de precio ----------------------------------------------------------

TIPO = "tipo_precio_recomendado"


def _excel_con_tipo(precio, tipo, cabecera: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["producto", "sucursal_id", "Precio Recomendado Compra", cabecera])
    ws.append(["17 GK2Z9365C", "LINDEROS", precio, tipo])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.parametrize("cabecera", [
    "tipo_precio_recomendado",  # como lo manda el motor
    "Tipo de Precio",           # como sale en el Excel que exporta la plataforma
])
def test_el_tipo_de_precio_llega_por_las_dos_cabeceras(db_session, cabecera):
    """El motor y el export no le dicen igual a la misma columna.

    Si solo se aceptara la del motor, volver a subir un Excel bajado de la propia
    plataforma perderia la columna, con un "Columnas ignoradas (sin mapeo)" que
    nadie lee.
    """
    excel_loader.cargar_sugerido(
        db_session, "sugerido.xlsx", _excel_con_tipo(32422, "Flota", cabecera))

    fila = db_session.query(Sugerido).filter_by(producto="17 GK2Z9365C").one()
    assert getattr(fila, CAMPO) == 32422
    assert getattr(fila, TIPO) == "Flota"


def test_la_migracion_repone_el_tipo(engine_con_tablas_viejas):
    eng = engine_con_tablas_viejas
    with eng.begin() as cx:
        cx.execute(text(f"ALTER TABLE sugerido DROP COLUMN {TIPO}"))

    create_all()

    assert TIPO in {c["name"] for c in inspect(eng).get_columns("sugerido")}


def test_el_export_incluye_el_tipo():
    assert TIPO in LABELS

    wb = load_workbook(io.BytesIO(generar_excel(
        [{"producto": "A", CAMPO: 32422, TIPO: "Flota"}], ["producto", CAMPO, TIPO])))
    hoja = wb.active

    assert [c.value for c in hoja[1]] == ["Producto", "Precio Recomendado Compra", "Tipo de Precio"]
    assert hoja.cell(row=2, column=3).value == "Flota"


def test_el_tipo_esta_en_el_catalogo_del_frontend():
    from pathlib import Path

    columnas = Path(__file__).resolve().parents[2] / "web" / "lib" / "columnas.ts"
    assert f'key: "{TIPO}"' in columnas.read_text(encoding="utf-8")
