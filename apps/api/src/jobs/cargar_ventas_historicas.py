"""Carga el historico de ventas (desde 2018) a la plataforma.

Lee los respaldos anuales de la CARPETA OFICIAL de datos, los agrega por
(periodo, producto, sucursal) y los sube. Se corre a mano cuando se agrega o
corrige un respaldo; no es parte de la sync diaria (el historico no cambia).

    python -m src.jobs.cargar_ventas_historicas                 # todos los anos
    python -m src.jobs.cargar_ventas_historicas --anio 2026     # solo uno

Carpeta: variable VENTAS_HISTORICAS_DIR, o la subcarpeta "Ventas" de
MOTOR_CRUDOS_DIR. Cada archivo se identifica por el ano que trae en el nombre.
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from sqlalchemy import delete, insert

from ..db import SessionLocal, create_all
from ..models import VentaHistorica

# Columnas que se leen del respaldo (el resto no se usa y pesa).
COL_PERIODO, COL_PRODUCTO, COL_SUCURSAL = "Periodo", "Producto", "SUCURSAL"
COL_CANTIDAD, COL_NETO = "Cantidad", "Total Neta"
CHUNK = 1000


def _carpeta() -> Path:
    d = os.environ.get("VENTAS_HISTORICAS_DIR")
    if d:
        return Path(d)
    crudos = os.environ.get("MOTOR_CRUDOS_DIR")
    if crudos:
        sub = Path(crudos) / "Ventas"
        return sub if sub.exists() else Path(crudos)
    raise RuntimeError(
        "Define VENTAS_HISTORICAS_DIR (o MOTOR_CRUDOS_DIR) con la carpeta de respaldos."
    )


def _num(v) -> float:
    """Numero tolerante: el respaldo trae textos, guiones y celdas vacias."""
    if v is None or isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(".", "").replace(",", ".").strip())
    except ValueError:
        return 0.0


def _agregar(ruta: Path) -> dict[tuple[str, str, str | None], list]:
    """Agrega un respaldo por (periodo, producto, sucursal).

    Se usa openpyxl en modo read_only y no polars: este job vive en el repo de la
    plataforma, que se despliega en la nube, y no vale la pena sumarle una
    dependencia pesada por un proceso que se corre a mano un par de veces al ano.
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        cabecera = [str(c).strip() if c is not None else "" for c in next(it)]
        faltan = {COL_PERIODO, COL_PRODUCTO, COL_CANTIDAD} - set(cabecera)
        if faltan:
            raise ValueError(f"{ruta.name}: faltan columnas {sorted(faltan)}")
        i_per, i_prod = cabecera.index(COL_PERIODO), cabecera.index(COL_PRODUCTO)
        i_cant = cabecera.index(COL_CANTIDAD)
        i_suc = cabecera.index(COL_SUCURSAL) if COL_SUCURSAL in cabecera else None
        i_neto = cabecera.index(COL_NETO) if COL_NETO in cabecera else None

        acum: dict[tuple[str, str, str | None], list] = {}
        for row in it:
            if not row or row[i_per] is None or row[i_prod] is None:
                continue
            clave = (
                str(row[i_per]).strip(),
                str(row[i_prod]).strip(),
                str(row[i_suc]).strip() if i_suc is not None and row[i_suc] else None,
            )
            reg = acum.setdefault(clave, [0.0, 0.0, 0])
            reg[0] += _num(row[i_cant])
            if i_neto is not None:
                reg[1] += _num(row[i_neto])
            reg[2] += 1
        return acum
    finally:
        wb.close()


def cargar(db, ruta: Path) -> int:
    """Carga un respaldo. Reemplaza SOLO los periodos que ese archivo trae, para
    poder recargar un ano corregido sin borrar el resto del historico."""
    acum = _agregar(ruta)
    periodos = {p for (p, _, _) in acum}
    if not periodos:
        return 0
    db.execute(delete(VentaHistorica).where(VentaHistorica.periodo.in_(periodos)))
    registros = [
        {
            "tenant_id": "curifor",
            "periodo": per,
            "producto": prod,
            "sucursal": suc,
            "cantidad": vals[0],
            "neto": vals[1] or None,
            "n_lineas": vals[2],
        }
        for (per, prod, suc), vals in acum.items()
    ]
    for i in range(0, len(registros), CHUNK):
        db.execute(insert(VentaHistorica).values(registros[i : i + CHUNK]))
    db.commit()
    return len(registros)


def run(anio: str | None = None) -> int:
    carpeta = _carpeta()
    archivos = sorted(
        p for p in carpeta.glob("*.xlsx")
        if not p.name.startswith("~$") and (anio is None or anio in p.name)
    )
    if not archivos:
        print(f"No hay respaldos en {carpeta}", file=sys.stderr)
        return 1
    create_all()
    db = SessionLocal()
    total = 0
    try:
        for p in archivos:
            print(f"  {p.name} ...", flush=True)
            n = cargar(db, p)
            total += n
            print(f"  {p.name}: {n:,} filas")
    finally:
        db.close()
    print(f"TOTAL cargado: {total:,} filas")
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Carga el historico de ventas.")
    ap.add_argument("--anio", help="Cargar solo los archivos que contengan este ano")
    args = ap.parse_args()
    raise SystemExit(run(anio=args.anio))
