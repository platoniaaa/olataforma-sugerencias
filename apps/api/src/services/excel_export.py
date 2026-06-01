"""Genera un archivo Excel (.xlsx) del sugerido filtrado, con las columnas elegidas.

Usa openpyxl. Aplica formato chileno: CLP sin decimales con miles en punto.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Nota: los rows vienen como dicts desde sugerido_service.listar (mezcla
# sugerido + catalogo + manuales). Antes eran ORM y este modulo hacia getattr,
# pero ahora son dicts.

# Etiquetas legibles para las cabeceras del Excel.
LABELS: dict[str, str] = {
    "producto": "Producto",
    "descripcion": "Descripcion",
    "clasificacion_abc": "ABC",
    "nombre_sucursal": "Sucursal",
    "sucursal_id": "ID Sucursal",
    "proveedor": "Proveedor",
    "filtro1_final": "Marca",
    "tipo_origen": "Tipo Origen",
    "unidad_medida": "Unidad",
    "lead_time_dias": "Lead Time (dias)",
    "lt_efectivo": "LT Efectivo",
    "lt_origen": "Origen LT",
    "abastece_cd": "Abastece CD",
    "prioridad_cd": "Prioridad CD",
    "demanda_mensual": "Demanda Mensual",
    "demanda_diaria": "Demanda Diaria",
    "desv_std_mensual": "Desv Std Mensual",
    "stock_seguridad": "Stock Seguridad",
    "punto_de_pedido": "Punto de Pedido",
    "costo_unitario": "Costo Unitario",
    "pedir": "Pedir",
    "stock_activo_suc": "Stock Activo",
    "stock_en_transito_suc": "Stock en Transito",
    "stock_en_cd": "Stock en CD",
    "sugerido_traslado": "Sugerido Traslado",
    "sugerido_compra_neto": "Sugerido Compra Neto",
    "total_sugerido_suc": "Total Sugerido",
    "total_valor_sugerido_clp": "Valor Total CLP",
}

# Columnas por defecto si el cliente no especifica.
DEFAULT_COLUMNS = [
    "producto", "descripcion", "clasificacion_abc", "nombre_sucursal",
    "proveedor", "total_sugerido_suc", "total_valor_sugerido_clp",
]

CLP_COLUMNS = {"total_valor_sugerido_clp", "costo_unitario"}
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generar_excel(rows: list[dict], columnas: list[str] | None) -> bytes:
    cols = [c for c in (columnas or []) if c in LABELS] or DEFAULT_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "Sugerido"

    # Cabecera.
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=LABELS.get(col, col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos. rows son dicts (sugerido + catalogo + manuales mezclados).
    # Fallback a getattr por si alguien pasa ORM, pero el caller actual usa dicts.
    def _valor(row, col):
        if isinstance(row, dict):
            return row.get(col)
        return getattr(row, col, None)

    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            value = _valor(row, col)
            cell = ws.cell(row=i, column=j, value=value)
            if col in CLP_COLUMNS and isinstance(value, (int, float)):
                cell.number_format = '"$"#,##0'
            elif isinstance(value, float):
                cell.number_format = "#,##0.00"

    # Ancho de columnas aproximado.
    for j, col in enumerate(cols, start=1):
        width = max(12, min(40, len(LABELS.get(col, col)) + 4))
        ws.column_dimensions[get_column_letter(j)].width = width

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo() -> str:
    return f"sugerido_{date.today():%Y%m%d}.xlsx"


def _sanear_hoja(nombre: str) -> str:
    # Excel: max 31 chars y sin estos caracteres.
    for ch in r"[]:*?/\\":
        nombre = nombre.replace(ch, " ")
    return nombre[:31] or "Proveedor"


def generar_orden_compra(carros: list, proveedor: str | None = None) -> bytes:
    """Genera un Excel con la orden de compra: una hoja por proveedor.

    `carros` es una lista de CarroProveedor (schema). Si `proveedor` viene, solo ese.
    """
    seleccion = [c for c in carros if proveedor is None or c.proveedor == proveedor]
    wb = Workbook()
    wb.remove(wb.active)

    cols = [
        ("Producto", "producto", "texto"),
        ("Descripcion", "descripcion", "texto"),
        ("ABC", "clasificacion_abc", "texto"),
        ("Cantidad", "cantidad", "numero"),
        ("Costo Unitario", "costo_unitario", "clp"),
        ("Subtotal", "subtotal_clp", "clp"),
    ]
    nombres_usados: set[str] = set()
    for carro in seleccion:
        base = _sanear_hoja(carro.proveedor)
        nombre = base
        i = 2
        while nombre in nombres_usados:
            nombre = f"{base[:28]} {i}"
            i += 1
        nombres_usados.add(nombre)
        ws = wb.create_sheet(title=nombre)

        ws["A1"] = "ORDEN DE COMPRA"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Proveedor: {carro.proveedor}"
        ws["A3"] = f"Fecha: {date.today():%d-%m-%Y}"
        ws["A4"] = f"Productos: {carro.n_productos}   Total: ${carro.total_clp:,.0f}".replace(",", ".")
        for c in ("A2", "A3", "A4"):
            ws[c].font = Font(bold=True)

        fila_hdr = 6
        for j, (label, _key, _t) in enumerate(cols, start=1):
            cell = ws.cell(row=fila_hdr, column=j, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for i_row, linea in enumerate(carro.lineas, start=fila_hdr + 1):
            for j, (_label, key, tipo) in enumerate(cols, start=1):
                val = getattr(linea, key, None)
                cell = ws.cell(row=i_row, column=j, value=val)
                if tipo == "clp" and isinstance(val, (int, float)):
                    cell.number_format = '"$"#,##0'
                elif tipo == "numero" and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

        # Total al pie
        fila_total = fila_hdr + 1 + len(carro.lineas)
        ws.cell(row=fila_total, column=5, value="TOTAL").font = Font(bold=True)
        tc = ws.cell(row=fila_total, column=6, value=carro.total_clp)
        tc.number_format = '"$"#,##0'
        tc.font = Font(bold=True)

        anchos = [16, 40, 6, 12, 16, 16]
        for j, w in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = f"A{fila_hdr + 1}"

    if not wb.sheetnames:
        wb.create_sheet(title="Sin datos")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_orden(proveedor: str | None = None) -> str:
    if proveedor:
        slug = "".join(c for c in proveedor if c.isalnum() or c in " -_").strip()[:30]
        return f"orden_{slug}_{date.today():%Y%m%d}.xlsx"
    return f"ordenes_compra_{date.today():%Y%m%d}.xlsx"
