"""La lista de precios: calcularla, consultarla, editarla y exportarla al ERP.

Es la logica que vivia en las formulas de `LISTA DE PRECIOS.xlsx` y en el .exe
que corria en el PC de Hugo, ahora como codigo que corre donde estan los datos.

Como se decide el precio de un producto (`calcular`, en este orden, y gana el
primero que aplica):

  1. Precio fijo (override)      -> ese es el precio, cambie lo que cambie.
  2. Congelado (override)         -> el precio que tenia al congelar.
  3. No es producto (override)    -> sin precio (servicios, mano de obra).
  4. Stock 0 y nada en transito   -> precio 0 (el ERP no lo ofrece).
  5. Tipo Sugerido                -> la lista del proveedor (Gildemeister).
  6. El resto                     -> ROUND(costo x factor).

El factor sale de la politica por (tipo, procedencia). El tipo y la procedencia
se deciden en cascada; el detalle esta en `_tipo` y `_procedencia`.

La regla de la procedencia por compras es la que pidio Abastecimiento: si la
ultima compra esta en el seguimiento de importacion, es Importado; si esta en el
nacional, es Nacional; si esta en los dos, manda la mas reciente.
"""
from __future__ import annotations

import io
import math
import re
import uuid
from datetime import date, datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func, insert, or_, select, update
from sqlalchemy.orm import Session

from ..config import get_settings
from ..models import (
    DimProducto,
    PoliticaPrecio,
    PrecioCambio,
    PrecioEnvio,
    PrecioOverride,
    PrecioProducto,
    ProductoCatalogo,
    StockTransito,
    StockUnificado,
    Sugerido,
    VentaHistorica,
)
from . import auditoria_service, politica_precio_service as politica
from .politica_precio_service import IMPORTADO, NACIONAL, SIN_REVISION, TIPO_SUGERIDO

settings = get_settings()

_RUBRO = re.compile(r"^(\d{1,3})\s+")
_LOTE = 500          # filas por INSERT
_IN = 800            # codigos por clausula IN (SQLite acepta 999 parametros)

ESTADOS = ("OK", "FIJO", "CONGELADO", "SUGERIDO", "SIN REVISION", "NO PRODUCTO", "SIN STOCK")
CAMPOS_CAMBIO = ("procedencia", "costo", "stock", "precio", "tipo")


def _now() -> datetime:
    return datetime.now(timezone.utc)


def rubro_de(producto: str | None) -> str | None:
    m = _RUBRO.match((producto or "").strip())
    return m.group(1) if m else None


def _norm_proc(valor: str | None) -> str | None:
    v = (valor or "").strip().upper()
    if v in ("IMPORTADO", "IMPORTADOS"):
        return IMPORTADO
    if v in ("NACIONAL", "NACIONALES"):
        return NACIONAL
    return None


def _num(v) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def redondear(v: float) -> int:
    """Redondeo como el de Excel (mitad hacia arriba). El `round` de Python
    redondea la mitad al par: 124.464,5 da 124.464 y Excel da 124.465, y esa
    diferencia de $1 aparecia como "cambio de precio" en miles de productos."""
    # 54.115 x 2,3 da 124.464,49999999999 en binario; Excel lo ve como ,5 porque
    # trabaja a 15 digitos. Se recorta a 6 decimales antes de decidir.
    return int(math.floor(round(float(v), 6) + 0.5))


def _fecha(v) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


# ----------------------------------------------------------------- la regla
def _tipo(fila: dict, ov: dict | None, rubros: dict) -> tuple[str | None, str | None]:
    """(tipo, de donde salio). Manual > glosa que empieza con NEU > rubro."""
    if ov and ov.get("tipo_manual"):
        return ov["tipo_manual"], "manual"
    # Se exige al INICIO de la glosa para no agarrar "SENSOR PRESION NEUMATICO".
    if (fila.get("glosa") or "").strip().upper().startswith("NEU"):
        return "Neumatico", "glosa"
    r = rubros.get(fila.get("rubro") or "")
    if r and r.get("tipo"):
        return r["tipo"], "rubro"
    return None, None


def _procedencia(fila: dict, ov: dict | None, rubros: dict) -> tuple[str, str]:
    """(procedencia, de donde salio). Manual > rubro forzado > compras > maestro > SIN REVISION."""
    if ov and ov.get("procedencia_manual"):
        return ov["procedencia_manual"], "manual"
    r = rubros.get(fila.get("rubro") or "")
    if r and r.get("procedencia_forzada"):
        return r["procedencia_forzada"], "rubro"
    imp, nac = _fecha(fila.get("ult_recep_importado")), _fecha(fila.get("ult_pe_nacional"))
    if imp or nac:
        if imp and nac:
            return (IMPORTADO if imp >= nac else NACIONAL), "compras"
        return (IMPORTADO if imp else NACIONAL), "compras"
    m = _norm_proc(fila.get("procedencia_maestro"))
    if m:
        return m, "maestro"
    return SIN_REVISION, "default"


def calcular(fila: dict, ov: dict | None, factores: dict, rubros: dict) -> dict:
    """Aplica la regla a UNA fila. Pura: no toca la base, para poder probarla.

    `fila` trae: producto, glosa, rubro, procedencia_maestro, costo, stock,
    stock_transito, ult_recep_importado, ult_pe_nacional, precio_sugerido.
    Devuelve los campos calculados listos para escribir en `precio_producto`.
    """
    ov = ov or {}
    tipo, tipo_origen = _tipo(fila, ov, rubros)
    proc, proc_origen = _procedencia(fila, ov, rubros)
    factor = factores.get((politica.tipo_canonico(tipo), (proc or "").lower())) if tipo else None

    costo = _num(fila.get("costo")) or 0.0
    stock = _num(fila.get("stock")) or 0.0
    transito = _num(fila.get("stock_transito")) or 0.0
    sugerido = _num(fila.get("precio_sugerido"))

    # 3 a 6: lo que da la regla sin mirar precio fijo ni congelado. El orden es
    # el del .exe (LEEME): stock 0 manda antes que "no es producto", asi que un
    # servicio sin stock sale en 0 igual que hoy.
    if stock <= 0 and transito <= 0:
        calculado, estado = 0.0, "SIN STOCK"
    elif ov.get("no_producto"):
        calculado, estado = None, "NO PRODUCTO"
    elif (tipo or "").lower() == TIPO_SUGERIDO.lower():
        calculado = float(redondear(sugerido)) if sugerido and sugerido > 0 else None
        estado = "SUGERIDO" if calculado is not None else "SIN REVISION"
    elif factor and costo > 0:
        calculado, estado = float(redondear(costo * factor)), "OK"
    else:
        calculado, estado = None, "SIN REVISION"

    # 1 y 2: la gente le gana a la regla.
    if ov.get("precio_fijo") is not None:
        final, estado = float(ov["precio_fijo"]), "FIJO"
    elif ov.get("congelar"):
        congelado = ov.get("congelado_precio")
        final = float(congelado) if congelado is not None else calculado
        estado = "CONGELADO"
    else:
        final = calculado

    return {
        "tipo": tipo, "tipo_origen": tipo_origen,
        "procedencia_final": proc, "procedencia_origen": proc_origen,
        "factor": factor,
        "precio_calculado": calculado, "precio_final": final, "estado": estado,
    }


# ------------------------------------------------------------ insumos (foto)
def _en_lotes(codigos: list[str]):
    for i in range(0, len(codigos), _IN):
        yield codigos[i:i + _IN]


def _stock(db: Session, codigos: list[str]) -> tuple[dict[str, float], dict[str, float]]:
    stock: dict[str, float] = {}
    transito: dict[str, float] = {}
    for lote in _en_lotes(codigos):
        try:
            for p, t in db.execute(
                select(StockUnificado.producto, func.coalesce(func.sum(StockUnificado.stock), 0))
                .where(StockUnificado.producto.in_(lote)).group_by(StockUnificado.producto)
            ).all():
                stock[p] = float(t or 0)
            for p, t in db.execute(
                select(StockTransito.producto, func.coalesce(func.sum(StockTransito.cantidad), 0))
                .where(StockTransito.producto.in_(lote)).group_by(StockTransito.producto)
            ).all():
                transito[p] = float(t or 0)
        except Exception:  # noqa: BLE001 - tabla ausente
            db.rollback()
    return stock, transito


def _costos(db: Session, codigos: list[str]) -> dict[str, float]:
    """El costo mas fresco que tiene la plataforma: el del BI (dim_producto) y,
    si no esta, el del maestro del ERP (producto_catalogo)."""
    costo: dict[str, float] = {}
    for lote in _en_lotes(codigos):
        try:
            for p, c in db.execute(
                select(ProductoCatalogo.producto, ProductoCatalogo.costo)
                .where(ProductoCatalogo.producto.in_(lote))
            ).all():
                if c and c > 0:
                    costo[p] = float(c)
            for p, c in db.execute(
                select(DimProducto.producto, DimProducto.costo_unitario)
                .where(DimProducto.producto.in_(lote))
            ).all():
                if c and c > 0:
                    costo[p] = float(c)
        except Exception:  # noqa: BLE001
            db.rollback()
    return costo


def _sugeridos(db: Session, codigos: list[str]) -> dict[str, float]:
    """Precio de lista Gildemeister, que el motor publica pegado al sugerido."""
    out: dict[str, float] = {}
    for lote in _en_lotes(codigos):
        try:
            for p, v in db.execute(
                select(Sugerido.producto, func.max(Sugerido.precio_sugerido_gilde))
                .where(Sugerido.producto.in_(lote)).group_by(Sugerido.producto)
            ).all():
                if v and v > 0:
                    out[p] = float(v)
        except Exception:  # noqa: BLE001
            db.rollback()
    return out


def _ultima_venta(db: Session, codigos: list[str]) -> dict[str, date]:
    out: dict[str, date] = {}
    for lote in _en_lotes(codigos):
        try:
            for p, per in db.execute(
                select(VentaHistorica.producto, func.max(VentaHistorica.periodo))
                .where(VentaHistorica.producto.in_(lote), VentaHistorica.cantidad > 0)
                .group_by(VentaHistorica.producto)
            ).all():
                if per and len(per) == 6 and per.isdigit():
                    out[p] = date(int(per[:4]), int(per[4:]), 1)
        except Exception:  # noqa: BLE001
            db.rollback()
    return out


def _overrides(db: Session, codigos: list[str] | None = None) -> dict[str, dict]:
    stmt = select(PrecioOverride).where(PrecioOverride.tenant_id == settings.default_tenant_id)
    out: dict[str, dict] = {}
    try:
        if codigos is None:
            filas = db.scalars(stmt).all()
        else:
            filas = []
            for lote in _en_lotes(codigos):
                filas += db.scalars(stmt.where(PrecioOverride.producto.in_(lote))).all()
    except Exception:  # noqa: BLE001
        db.rollback()
        return {}
    for o in filas:
        out[o.producto] = _override_dict(o)
    return out


def _override_dict(o: PrecioOverride) -> dict:
    return {
        "precio_fijo": o.precio_fijo, "congelar": bool(o.congelar),
        "congelado_precio": o.congelado_precio,
        "congelado_en": o.congelado_en.isoformat() if o.congelado_en else None,
        "tipo_manual": o.tipo_manual, "procedencia_manual": o.procedencia_manual,
        "no_producto": bool(o.no_producto), "obs": o.obs,
        "editado_por": o.editado_por,
        "editado_en": o.editado_en.isoformat() if o.editado_en else None,
    }


# ----------------------------------------------------------------- recalculo
def _txt(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, float):
        return f"{v:.2f}".rstrip("0").rstrip(".")
    return str(v)


def recalcular(db: Session, usuario: str | None = None, refrescar_insumos: bool = True) -> dict:
    """Recalcula TODA la lista y anota que cambio. Es idempotente: correrlo dos
    veces seguidas no genera cambios la segunda vez.

    `refrescar_insumos=False` recalcula solo con lo que la tabla ya tiene (util
    tras editar una politica, sin volver a leer stock y costo)."""
    tenant = settings.default_tenant_id
    filas = db.scalars(select(PrecioProducto).where(PrecioProducto.tenant_id == tenant)).all()
    if not filas:
        return {"productos": 0, "cambios": 0, "corrida_id": None}
    codigos = [f.producto for f in filas]
    fact = politica.factores(db)
    rub = politica.rubros(db)
    ovs = _overrides(db)
    if refrescar_insumos:
        stock, transito = _stock(db, codigos)
        costos = _costos(db, codigos)
        sugeridos = _sugeridos(db, codigos)
        ventas = _ultima_venta(db, codigos)
    else:
        stock = transito = costos = sugeridos = ventas = {}

    corrida = str(uuid.uuid4())
    ahora = _now()
    cambios: list[dict] = []
    por_campo = {c: 0 for c in CAMPOS_CAMBIO}
    pendientes: dict[str, int] = {}

    for f in filas:
        # La foto ANTES de tocar nada: contra esto se miden los cambios. Si se
        # tomara despues de refrescar el stock y el costo, esos dos nunca
        # aparecerian como cambio.
        antes = {
            "procedencia": f.procedencia_final, "costo": f.costo, "stock": f.stock,
            "precio": f.precio_final, "tipo": f.tipo,
        }
        if refrescar_insumos:
            # Los insumos solo se pisan cuando la plataforma tiene el dato; si no,
            # se conserva la ultima foto (la del Excel en la primera carga).
            if f.producto in stock or f.producto in transito:
                f.stock = stock.get(f.producto, 0.0)
                f.stock_transito = transito.get(f.producto, 0.0)
            if f.producto in costos:
                f.costo = costos[f.producto]
            if f.producto in sugeridos:
                f.precio_sugerido = sugeridos[f.producto]
            if f.producto in ventas:
                f.ultima_venta = ventas[f.producto]
        if not f.rubro:
            f.rubro = rubro_de(f.producto)

        r = calcular(
            {
                "glosa": f.glosa, "rubro": f.rubro, "procedencia_maestro": f.procedencia_maestro,
                "costo": f.costo, "stock": f.stock, "stock_transito": f.stock_transito,
                "ult_recep_importado": f.ult_recep_importado, "ult_pe_nacional": f.ult_pe_nacional,
                "precio_sugerido": f.precio_sugerido,
            },
            ovs.get(f.producto), fact, rub,
        )
        for k, v in r.items():
            setattr(f, k, v)
        f.actualizado_en = ahora

        despues = {
            "procedencia": f.procedencia_final, "costo": f.costo, "stock": f.stock,
            "precio": f.precio_final, "tipo": f.tipo,
        }
        # Solo cuenta como cambio lo que YA tenia un valor y ahora tiene otro. En la
        # primera corrida todo pasa de vacio a algo y eso no es un cambio que
        # alguien tenga que revisar.
        for campo in CAMPOS_CAMBIO:
            a, d = antes[campo], despues[campo]
            if a is None or _txt(a) == _txt(d):
                continue
            cambios.append({
                "tenant_id": tenant, "producto": f.producto, "campo": campo,
                "antes": _txt(a), "despues": _txt(d), "corrida_id": corrida,
                "detectado_en": ahora, "visto": False,
            })
            por_campo[campo] += 1
            pendientes[f.producto] = pendientes.get(f.producto, 0) + 1

    for i in range(0, len(cambios), _LOTE):
        db.execute(insert(PrecioCambio), cambios[i:i + _LOTE])
    for f in filas:
        if f.producto in pendientes:
            f.cambios_pendientes = (f.cambios_pendientes or 0) + pendientes[f.producto]

    if cambios:
        detalle = ", ".join(f"{k}: {v}" for k, v in por_campo.items() if v)
        auditoria_service.notificar(
            db, tipo="precios",
            titulo=f"Lista de precios recalculada: {len(cambios)} cambios en {len(pendientes)} productos",
            mensaje=detalle, creado_por_email=usuario,
        )
    auditoria_service.registrar(
        db, accion="precios_recalculados", entidad="precios", usuario_email=usuario,
        detalle=f"{len(filas)} productos, {len(cambios)} cambios",
    )
    db.commit()
    return {
        "productos": len(filas), "cambios": len(cambios),
        "productos_con_cambios": len(pendientes), "por_campo": por_campo,
        "corrida_id": corrida,
    }


# ----------------------------------------------------------------- consultas
_ORDENABLES = {
    "producto", "glosa", "rubro", "tipo", "procedencia_final", "costo", "precio_erp",
    "stock", "precio_final", "precio_calculado", "estado", "ultima_venta",
    "cambios_pendientes", "factor", "actualizado_en",
}


def _filtrar(stmt, f):
    tenant = settings.default_tenant_id
    stmt = stmt.where(PrecioProducto.tenant_id == tenant)
    q = (getattr(f, "q", None) or "").strip()
    if q:
        like = f"%{q}%"
        stmt = stmt.where(or_(PrecioProducto.producto.ilike(like), PrecioProducto.glosa.ilike(like)))
    if getattr(f, "rubro", None):
        stmt = stmt.where(PrecioProducto.rubro.in_(f.rubro))
    if getattr(f, "tipo", None):
        stmt = stmt.where(PrecioProducto.tipo.in_(f.tipo))
    if getattr(f, "procedencia", None):
        stmt = stmt.where(PrecioProducto.procedencia_final.in_(f.procedencia))
    if getattr(f, "estado", None):
        stmt = stmt.where(PrecioProducto.estado.in_(f.estado))
    if getattr(f, "origen", None):
        stmt = stmt.where(PrecioProducto.origen == f.origen)
    if getattr(f, "con_cambios", False):
        stmt = stmt.where(PrecioProducto.cambios_pendientes > 0)
    if getattr(f, "con_stock", False):
        stmt = stmt.where(PrecioProducto.stock > 0)
    return stmt


def listar(db: Session, f, page: int = 1, limit: int = 200, sort: str | None = None):
    base = _filtrar(select(PrecioProducto), f)
    total = db.scalar(select(func.count()).select_from(base.subquery())) or 0
    col, desc_ = "producto", False
    if sort:
        col = sort.lstrip("-")
        desc_ = sort.startswith("-")
        if col not in _ORDENABLES:
            col, desc_ = "producto", False
    c = getattr(PrecioProducto, col)
    stmt = base.order_by(c.desc().nulls_last() if desc_ else c.asc().nulls_last(), PrecioProducto.producto)
    items = list(db.scalars(stmt.offset((page - 1) * limit).limit(limit)).all())
    ovs = _overrides(db, [i.producto for i in items]) if items else {}
    return [_fila_dict(i, ovs.get(i.producto)) for i in items], total


def _fila_dict(p: PrecioProducto, ov: dict | None) -> dict:
    d = {c.name: getattr(p, c.name) for c in PrecioProducto.__table__.columns}
    for k in ("creado_en", "actualizado_en"):
        d[k] = d[k].isoformat() if d.get(k) else None
    for k in ("ult_recep_importado", "ult_pe_nacional", "ultima_venta"):
        d[k] = d[k].isoformat() if d.get(k) else None
    d["desviacion_pesos"] = (
        (p.precio_erp - p.precio_final)
        if p.precio_erp is not None and p.precio_final is not None and p.precio_erp > 0 else None
    )
    d["desviacion_pct"] = (
        round((p.precio_erp - p.precio_final) / p.precio_final * 100, 1)
        if p.precio_erp and p.precio_final and p.precio_final > 0 else None
    )
    ov = ov or {}
    d["precio_fijo"] = ov.get("precio_fijo")
    d["congelar"] = bool(ov.get("congelar"))
    d["congelado_precio"] = ov.get("congelado_precio")
    d["no_producto"] = bool(ov.get("no_producto"))
    d["obs"] = ov.get("obs")
    d["tipo_manual"] = ov.get("tipo_manual")
    d["procedencia_manual"] = ov.get("procedencia_manual")
    d["editado_por"] = ov.get("editado_por")
    d["editado_en"] = ov.get("editado_en")
    return d


def opciones_filtros(db: Session) -> dict:
    tenant = settings.default_tenant_id

    def _distintos(col):
        try:
            vals = db.execute(
                select(col).where(PrecioProducto.tenant_id == tenant, col.isnot(None)).distinct()
            ).scalars().all()
        except Exception:  # noqa: BLE001
            db.rollback()
            return []
        return sorted(v for v in vals if v not in (None, ""))

    rubros_ = _distintos(PrecioProducto.rubro)
    rubros_ = sorted(rubros_, key=lambda r: (0, int(r)) if r.isdigit() else (1, r))
    return {
        "rubros": rubros_,
        "tipos": _distintos(PrecioProducto.tipo),
        "procedencias": _distintos(PrecioProducto.procedencia_final),
        "estados": [e for e in ESTADOS if e in set(_distintos(PrecioProducto.estado))],
    }


def resumen(db: Session) -> dict:
    tenant = settings.default_tenant_id
    try:
        total = db.scalar(select(func.count()).select_from(PrecioProducto).where(PrecioProducto.tenant_id == tenant)) or 0
        con_cambios = db.scalar(
            select(func.count()).select_from(PrecioProducto)
            .where(PrecioProducto.tenant_id == tenant, PrecioProducto.cambios_pendientes > 0)
        ) or 0
        por_estado = dict(db.execute(
            select(PrecioProducto.estado, func.count()).where(PrecioProducto.tenant_id == tenant)
            .group_by(PrecioProducto.estado)
        ).all())
        ultima = db.scalar(select(func.max(PrecioProducto.actualizado_en)).where(PrecioProducto.tenant_id == tenant))
        ultimo_envio = db.scalar(select(func.max(PrecioEnvio.enviado_en)).where(PrecioEnvio.tenant_id == tenant))
        overrides = db.scalar(select(func.count()).select_from(PrecioOverride).where(PrecioOverride.tenant_id == tenant)) or 0
    except Exception:  # noqa: BLE001
        db.rollback()
        return {"productos": 0, "con_cambios": 0, "por_estado": {}, "ultimo_recalculo": None,
                "ultimo_envio": None, "overrides": 0, "pendientes_envio": 0}
    return {
        "productos": total, "con_cambios": con_cambios,
        "por_estado": {k or "(sin estado)": v for k, v in por_estado.items()},
        "ultimo_recalculo": ultima.isoformat() if ultima else None,
        "ultimo_envio": ultimo_envio.isoformat() if ultimo_envio else None,
        "overrides": overrides,
        "pendientes_envio": len(_diferencias(db)),
    }


def detalle(db: Session, producto: str) -> dict | None:
    p = db.scalars(select(PrecioProducto).where(
        PrecioProducto.tenant_id == settings.default_tenant_id, PrecioProducto.producto == producto,
    )).first()
    if not p:
        return None
    d = _fila_dict(p, _overrides(db, [producto]).get(producto))
    cambios = db.scalars(
        select(PrecioCambio).where(
            PrecioCambio.tenant_id == settings.default_tenant_id, PrecioCambio.producto == producto,
        ).order_by(PrecioCambio.detectado_en.desc()).limit(50)
    ).all()
    d["cambios"] = [
        {"campo": c.campo, "antes": c.antes, "despues": c.despues, "visto": c.visto,
         "detectado_en": c.detectado_en.isoformat() if c.detectado_en else None}
        for c in cambios
    ]
    envios = db.scalars(
        select(PrecioEnvio).where(
            PrecioEnvio.tenant_id == settings.default_tenant_id, PrecioEnvio.producto == producto,
        ).order_by(PrecioEnvio.enviado_en.desc()).limit(10)
    ).all()
    d["envios"] = [
        {"precio": e.precio, "costo": e.costo, "lote_id": e.lote_id,
         "enviado_en": e.enviado_en.isoformat() if e.enviado_en else None, "enviado_por": e.enviado_por}
        for e in envios
    ]
    return d


# ------------------------------------------------------------------ escritura
def guardar_override(db: Session, producto: str, datos: dict, usuario: str | None) -> dict:
    """Crea o actualiza la decision humana sobre un producto y recalcula SOLO esa fila."""
    tenant = settings.default_tenant_id
    p = db.scalars(select(PrecioProducto).where(
        PrecioProducto.tenant_id == tenant, PrecioProducto.producto == producto,
    )).first()
    if not p:
        raise LookupError(f"{producto} no esta en la lista de precios")
    o = db.scalars(select(PrecioOverride).where(
        PrecioOverride.tenant_id == tenant, PrecioOverride.producto == producto,
    )).first()
    if o is None:
        o = PrecioOverride(tenant_id=tenant, producto=producto)
        db.add(o)

    detalle_ = []
    if "precio_fijo" in datos:
        v = datos["precio_fijo"]
        if v is not None and float(v) < 0:
            raise ValueError("El precio fijo no puede ser negativo")
        if (o.precio_fijo or None) != (v if v is None else float(v)):
            detalle_.append(f"precio fijo {o.precio_fijo} -> {v}")
        o.precio_fijo = None if v is None else float(v)
    if "congelar" in datos:
        nuevo = bool(datos["congelar"])
        if nuevo and not o.congelar:
            # Se congela con el precio VIGENTE. Es la foto que despues se respeta.
            o.congelado_precio = p.precio_final
            o.congelado_en = _now()
            detalle_.append(f"congelado en {p.precio_final}")
        elif not nuevo and o.congelar:
            o.congelado_precio = None
            o.congelado_en = None
            detalle_.append("descongelado")
        o.congelar = nuevo
    for campo in ("tipo_manual", "procedencia_manual"):
        if campo in datos:
            v = (datos[campo] or "").strip() or None
            if campo == "procedencia_manual" and v and v not in (NACIONAL, IMPORTADO):
                raise ValueError("La procedencia manual debe ser Nacional o Importado")
            if getattr(o, campo) != v:
                detalle_.append(f"{campo} {getattr(o, campo)} -> {v}")
            setattr(o, campo, v)
    if "no_producto" in datos:
        v = bool(datos["no_producto"])
        if o.no_producto != v:
            detalle_.append("no es producto" if v else "vuelve a ser producto")
        o.no_producto = v
    if "obs" in datos:
        o.obs = (datos["obs"] or "").strip() or None
    o.editado_por = usuario
    o.editado_en = _now()
    db.flush()

    _recalcular_una(db, p, _override_dict(o))
    auditoria_service.registrar(
        db, accion="precio_override", entidad="precio", entidad_id=producto, producto=producto,
        usuario_email=usuario, detalle="; ".join(detalle_) or "sin cambios", motivo=o.obs,
    )
    db.commit()
    return _fila_dict(p, _override_dict(o))


def quitar_override(db: Session, producto: str, usuario: str | None) -> dict:
    tenant = settings.default_tenant_id
    p = db.scalars(select(PrecioProducto).where(
        PrecioProducto.tenant_id == tenant, PrecioProducto.producto == producto,
    )).first()
    if not p:
        raise LookupError(f"{producto} no esta en la lista de precios")
    db.execute(delete(PrecioOverride).where(
        PrecioOverride.tenant_id == tenant, PrecioOverride.producto == producto,
    ))
    _recalcular_una(db, p, None)
    auditoria_service.registrar(
        db, accion="precio_override_quitado", entidad="precio", entidad_id=producto,
        producto=producto, usuario_email=usuario,
    )
    db.commit()
    return _fila_dict(p, None)


def _recalcular_una(db: Session, p: PrecioProducto, ov: dict | None) -> None:
    r = calcular(
        {
            "glosa": p.glosa, "rubro": p.rubro or rubro_de(p.producto),
            "procedencia_maestro": p.procedencia_maestro, "costo": p.costo, "stock": p.stock,
            "stock_transito": p.stock_transito, "ult_recep_importado": p.ult_recep_importado,
            "ult_pe_nacional": p.ult_pe_nacional, "precio_sugerido": p.precio_sugerido,
        },
        ov, politica.factores(db), politica.rubros(db),
    )
    for k, v in r.items():
        setattr(p, k, v)
    p.actualizado_en = _now()


def crear_producto(db: Session, datos: dict, usuario: str | None) -> dict:
    """Un producto nuevo desde la plataforma (origen manual). Queda listo para
    exportarse al ERP en el proximo envio."""
    tenant = settings.default_tenant_id
    producto = (datos.get("producto") or "").strip()
    if not producto:
        raise ValueError("Falta el codigo del producto")
    if db.scalar(select(PrecioProducto.id).where(
        PrecioProducto.tenant_id == tenant, PrecioProducto.producto == producto,
    )):
        raise ValueError(f"{producto} ya esta en la lista")
    p = PrecioProducto(
        tenant_id=tenant, producto=producto, glosa=(datos.get("glosa") or "").strip() or None,
        rubro=(datos.get("rubro") or "").strip() or rubro_de(producto),
        procedencia_maestro=(datos.get("procedencia") or "").strip() or None,
        costo=_num(datos.get("costo")), stock=_num(datos.get("stock")) or 0.0,
        stock_transito=0.0, origen="manual", creado_por=usuario,
    )
    db.add(p)
    db.flush()
    ov = None
    if datos.get("precio_fijo") is not None or datos.get("tipo"):
        o = PrecioOverride(
            tenant_id=tenant, producto=producto,
            precio_fijo=_num(datos.get("precio_fijo")),
            tipo_manual=(datos.get("tipo") or "").strip() or None,
            obs=(datos.get("obs") or "").strip() or None, editado_por=usuario,
        )
        db.add(o)
        db.flush()
        ov = _override_dict(o)
    _recalcular_una(db, p, ov)
    auditoria_service.registrar(
        db, accion="precio_producto_creado", entidad="precio", entidad_id=producto,
        producto=producto, usuario_email=usuario, detalle=p.glosa, motivo=datos.get("obs"),
    )
    db.commit()
    return _fila_dict(p, ov)


def marcar_vistos(db: Session, productos: list[str] | None, usuario: str | None) -> dict:
    tenant = settings.default_tenant_id
    stmt = update(PrecioCambio).where(PrecioCambio.tenant_id == tenant, PrecioCambio.visto.is_(False))
    stmt2 = update(PrecioProducto).where(PrecioProducto.tenant_id == tenant)
    if productos:
        stmt = stmt.where(PrecioCambio.producto.in_(productos))
        stmt2 = stmt2.where(PrecioProducto.producto.in_(productos))
    r = db.execute(stmt.values(visto=True, visto_por=usuario))
    db.execute(stmt2.values(cambios_pendientes=0))
    db.commit()
    return {"vistos": r.rowcount}


# --------------------------------------------------------------- exportacion
def _ultimo_envio(db: Session) -> dict[str, tuple[float | None, float | None]]:
    """{producto: (precio, costo)} del envio mas reciente de cada producto."""
    tenant = settings.default_tenant_id
    sub = (
        select(PrecioEnvio.producto, func.max(PrecioEnvio.enviado_en).label("m"))
        .where(PrecioEnvio.tenant_id == tenant).group_by(PrecioEnvio.producto).subquery()
    )
    try:
        filas = db.execute(
            select(PrecioEnvio.producto, PrecioEnvio.precio, PrecioEnvio.costo)
            .join(sub, (PrecioEnvio.producto == sub.c.producto) & (PrecioEnvio.enviado_en == sub.c.m))
            .where(PrecioEnvio.tenant_id == tenant)
        ).all()
    except Exception:  # noqa: BLE001
        db.rollback()
        return {}
    return {p: (pr, co) for p, pr, co in filas}


def _diferencias(db: Session) -> list[PrecioProducto]:
    """Productos cuyo precio o costo actual difiere del ultimo enviado (o nunca enviados)."""
    tenant = settings.default_tenant_id
    ultimo = _ultimo_envio(db)
    filas = db.scalars(select(PrecioProducto).where(PrecioProducto.tenant_id == tenant)).all()
    out = []
    for p in filas:
        if p.precio_final is None:
            continue
        env = ultimo.get(p.producto)
        if env is None:
            out.append(p)
            continue
        if _txt(env[0]) != _txt(p.precio_final) or _txt(env[1]) != _txt(redondear(p.costo or 0)):
            out.append(p)
    return out


def exportar(db: Session, *, solo_diferencias: bool, registrar: bool, usuario: str | None,
             formato: str = "erp") -> tuple[bytes, str, int]:
    """Excel para el ERP. `formato="erp"`: SKU | Precio_Optimo | Costo, igual que el
    .exe. `formato="completa"`: la lista con todas las columnas, para revisar.
    Con `registrar`, deja en `precio_envio` lo que salio, para que el proximo
    "solo diferencias" parta de aca."""
    tenant = settings.default_tenant_id
    if solo_diferencias:
        filas = _diferencias(db)
    else:
        filas = [p for p in db.scalars(select(PrecioProducto).where(PrecioProducto.tenant_id == tenant)).all()
                 if p.precio_final is not None]
    filas.sort(key=lambda p: p.producto)

    wb = Workbook()
    ws = wb.active
    fill = PatternFill("solid", fgColor="1F4E5F")
    font = Font(bold=True, color="FFFFFF")
    if formato == "erp":
        ws.title = "Precios"
        cab = ["SKU", "Precio_Optimo", "Costo"]
        for j, t in enumerate(cab, 1):
            c = ws.cell(row=1, column=j, value=t); c.fill = fill; c.font = font
        for i, p in enumerate(filas, 2):
            ws.cell(row=i, column=1, value=p.producto)
            ws.cell(row=i, column=2, value=redondear(p.precio_final)).number_format = "#,##0"
            ws.cell(row=i, column=3, value=redondear(p.costo or 0)).number_format = "#,##0"
        for col, w in zip("ABC", (24, 16, 16)):
            ws.column_dimensions[col].width = w
    else:
        ws.title = "Lista de precios"
        cols = [
            ("producto", "Producto"), ("glosa", "Glosa"), ("rubro", "Rubro"), ("tipo", "Tipo"),
            ("procedencia_final", "Procedencia"), ("factor", "Factor"), ("costo", "Costo"),
            ("precio_erp", "Precio ERP"), ("precio_calculado", "Precio calculado"),
            ("precio_final", "Precio final"), ("estado", "Estado"), ("stock", "Stock"),
            ("stock_transito", "En transito"), ("ultima_venta", "Ultima venta"),
        ]
        for j, (_, t) in enumerate(cols, 1):
            c = ws.cell(row=1, column=j, value=t); c.fill = fill; c.font = font
        for i, p in enumerate(filas, 2):
            for j, (k, _) in enumerate(cols, 1):
                v = getattr(p, k)
                c = ws.cell(row=i, column=j, value=v)
                if k in ("costo", "precio_erp", "precio_calculado", "precio_final") and isinstance(v, (int, float)):
                    c.number_format = "#,##0"
                elif isinstance(v, date):
                    c.number_format = "DD-MM-YYYY"
        for j, (k, t) in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = 40 if k == "glosa" else max(12, len(t) + 4)
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.alignment = Alignment(vertical="center")
    buf = io.BytesIO()
    wb.save(buf)

    if registrar and filas:
        lote = str(uuid.uuid4())
        ahora = _now()
        regs = [
            {"tenant_id": tenant, "producto": p.producto, "precio": p.precio_final,
             "costo": redondear(p.costo or 0), "lote_id": lote, "enviado_en": ahora, "enviado_por": usuario}
            for p in filas
        ]
        for i in range(0, len(regs), _LOTE):
            db.execute(insert(PrecioEnvio), regs[i:i + _LOTE])
        auditoria_service.registrar(
            db, accion="precios_exportados", entidad="precios", usuario_email=usuario,
            detalle=f"{len(filas)} productos ({'solo diferencias' if solo_diferencias else 'lista completa'})",
        )
        db.commit()
    nombre = f"precios_{'diferencias' if solo_diferencias else 'completa'}_{date.today():%Y%m%d}.xlsx"
    return buf.getvalue(), nombre, len(filas)


# -------------------------------------------------------------------- cargas
def cargar_maestro(db: Session, filas: list[dict], *, reemplazar: bool, usuario: str | None) -> dict:
    """Carga (o agrega a) la lista desde el Excel. Solo toca las filas `origen=maestro`.

    Cada fila: producto, glosa, rubro, tipo, procedencia_maestro, procedencia_final,
    costo, precio_erp, stock, stock_proyectado, obs_precio, precio_fijo, congelar,
    ultima_venta, ult_recep_importado, ult_pe_nacional.

    Lo humano del Excel (Obs Precio, Precio Fijo, Congelar) se convierte en
    overrides, pero SIN pisar un override que ya exista en la plataforma: si
    alguien ya decidio algo aca, eso vale mas que la foto del Excel.
    """
    tenant = settings.default_tenant_id
    if reemplazar:
        db.execute(delete(PrecioProducto).where(
            PrecioProducto.tenant_id == tenant, PrecioProducto.origen == "maestro",
        ))
        db.flush()
    existentes = {p for (p,) in db.execute(
        select(PrecioProducto.producto).where(PrecioProducto.tenant_id == tenant)
    ).all()}
    ovs_existentes = _overrides(db)

    ahora = _now()
    nuevos: list[dict] = []
    overrides: list[dict] = []
    ignoradas = 0
    fijo_sin_obs = 0
    for f in filas:
        producto = (f.get("producto") or "").strip()
        if not producto or producto in existentes:
            ignoradas += 1
            continue
        existentes.add(producto)
        nuevos.append({
            "tenant_id": tenant, "producto": producto,
            "glosa": (f.get("glosa") or "").strip() or None,
            "rubro": (str(f.get("rubro") or "").strip() or rubro_de(producto)),
            "tipo": (f.get("tipo") or "").strip() or None,
            "procedencia_maestro": (f.get("procedencia_maestro") or "").strip() or None,
            "procedencia_final": (f.get("procedencia_final") or "").strip() or None,
            "costo": _num(f.get("costo")), "precio_erp": _num(f.get("precio_erp")),
            "stock": _num(f.get("stock")) or 0.0,
            "stock_transito": _num(f.get("stock_proyectado")) or 0.0,
            "ult_recep_importado": _fecha(f.get("ult_recep_importado")),
            "ult_pe_nacional": _fecha(f.get("ult_pe_nacional")),
            "ultima_venta": _fecha(f.get("ultima_venta")),
            "precio_final": _num(f.get("precio_optimo_excel")),
            "origen": "maestro", "creado_en": ahora, "actualizado_en": ahora,
            "cambios_pendientes": 0,
        })
        obs = (f.get("obs_precio") or "").strip() or None
        fijo = _num(f.get("precio_fijo"))
        # La formula del Excel solo usa el Precio Fijo cuando hay una Obs escrita
        # (IF(AT="", calculo, AU)). Un fijo sin obs hoy no vale: se importa igual
        # que hoy, para no cambiar ningun precio por accidente. Se cuenta aparte.
        if fijo is not None and not obs:
            fijo_sin_obs += 1
            fijo = None
        congelar = str(f.get("congelar") or "").strip().lower() in ("1", "x", "si", "sí", "true")
        if (obs or fijo is not None or congelar) and producto not in ovs_existentes:
            overrides.append({
                "tenant_id": tenant, "producto": producto, "precio_fijo": fijo,
                "congelar": congelar, "congelado_precio": _num(f.get("precio_optimo_excel")) if congelar else None,
                "congelado_en": ahora if congelar else None, "obs": obs, "no_producto": False,
                "editado_por": usuario or "excel", "editado_en": ahora,
            })
    for i in range(0, len(nuevos), _LOTE):
        db.execute(insert(PrecioProducto), nuevos[i:i + _LOTE])
    for i in range(0, len(overrides), _LOTE):
        db.execute(insert(PrecioOverride), overrides[i:i + _LOTE])
    db.commit()
    return {"cargados": len(nuevos), "ignorados": ignoradas, "overrides": len(overrides),
            "precio_fijo_sin_obs_ignorado": fijo_sin_obs}


def conservar_clasificacion_excel(db: Session, usuario: str | None) -> dict:
    """Despues de la primera carga: donde el tipo o la procedencia del Excel NO
    coinciden con lo que la regla deduce, se guarda el valor del Excel como
    override manual. Es la forma de rescatar SOLO el trabajo hecho a mano, sin
    heredar como manual todo lo que una formula escribio."""
    tenant = settings.default_tenant_id
    fact = politica.factores(db)
    rub = politica.rubros(db)
    filas = db.scalars(select(PrecioProducto).where(PrecioProducto.tenant_id == tenant)).all()
    ovs = _overrides(db)
    n_tipo = n_proc = 0
    ahora = _now()
    for p in filas:
        base = {
            "glosa": p.glosa, "rubro": p.rubro or rubro_de(p.producto),
            "procedencia_maestro": p.procedencia_maestro, "ult_recep_importado": p.ult_recep_importado,
            "ult_pe_nacional": p.ult_pe_nacional,
        }
        tipo_regla, _ = _tipo(base, None, rub)
        proc_regla, _ = _procedencia(base, None, rub)
        tipo_x = (p.tipo or "").strip()
        proc_x = (p.procedencia_final or "").strip()
        quiere_tipo = tipo_x and politica.tipo_canonico(tipo_x) != politica.tipo_canonico(tipo_regla)
        quiere_proc = proc_x in (NACIONAL, IMPORTADO) and proc_x != proc_regla
        if not (quiere_tipo or quiere_proc):
            continue
        o = db.scalars(select(PrecioOverride).where(
            PrecioOverride.tenant_id == tenant, PrecioOverride.producto == p.producto,
        )).first()
        if o is None:
            o = PrecioOverride(tenant_id=tenant, producto=p.producto, editado_por=usuario or "excel", editado_en=ahora)
            db.add(o)
        if quiere_tipo and not o.tipo_manual:
            o.tipo_manual = tipo_x
            n_tipo += 1
        if quiere_proc and not o.procedencia_manual:
            o.procedencia_manual = proc_x
            n_proc += 1
        ovs[p.producto] = _override_dict(o)
    db.commit()
    return {"tipo_manual": n_tipo, "procedencia_manual": n_proc}


def cargar_compras(db: Session, filas: list[dict]) -> dict:
    """El agente publica la ultima compra por producto (seguimientos de compra).

    Cada fila: producto, ult_recep_importado?, ult_pe_nacional?. Es la regla de
    procedencia: se guarda la fecha en la columna que corresponde y el proximo
    recalculo decide."""
    tenant = settings.default_tenant_id
    n = 0
    for f in filas:
        producto = (f.get("producto") or "").strip()
        if not producto:
            continue
        vals = {}
        imp, nac = _fecha(f.get("ult_recep_importado")), _fecha(f.get("ult_pe_nacional"))
        if imp:
            vals["ult_recep_importado"] = imp
        if nac:
            vals["ult_pe_nacional"] = nac
        if not vals:
            continue
        r = db.execute(update(PrecioProducto).where(
            PrecioProducto.tenant_id == tenant, PrecioProducto.producto == producto,
        ).values(**vals))
        n += r.rowcount
    db.commit()
    return {"actualizados": n}


def cargar_no_productos(db: Session, productos: list[str], usuario: str | None) -> dict:
    tenant = settings.default_tenant_id
    n = 0
    for producto in {(p or "").strip() for p in productos if (p or "").strip()}:
        o = db.scalars(select(PrecioOverride).where(
            PrecioOverride.tenant_id == tenant, PrecioOverride.producto == producto,
        )).first()
        if o is None:
            o = PrecioOverride(tenant_id=tenant, producto=producto, editado_por=usuario or "excel")
            db.add(o)
        if not o.no_producto:
            o.no_producto = True
            n += 1
    db.commit()
    return {"marcados": n}


def cargar_precios_sugeridos(db: Session, filas: list[dict]) -> dict:
    """La lista de precios del proveedor (Gildemeister) para los tipo Sugerido.

    Cada fila: producto, precio_sin_iva (y opcionalmente precio_con_iva). Se usa
    el neto, que es como se habla de precios en la lista. Cuando el motor
    publique el precio pegado al sugerido, ese manda; esta carga es la foto
    inicial y el respaldo para lo que el sugerido no evalua."""
    tenant = settings.default_tenant_id
    n = 0
    for f in filas:
        producto = (f.get("producto") or "").strip()
        precio = _num(f.get("precio_sin_iva"))
        if not producto or not precio or precio <= 0:
            continue
        r = db.execute(update(PrecioProducto).where(
            PrecioProducto.tenant_id == tenant, PrecioProducto.producto == producto,
        ).values(precio_sugerido=precio))
        n += r.rowcount
    db.commit()
    return {"actualizados": n}
