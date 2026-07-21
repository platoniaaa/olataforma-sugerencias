"""Carga del Excel/CSV exportado del Power BI hacia la tabla `sugerido`.

Tolerante a las cabeceras: normaliza (minusculas, sin acentos, sin espacios) y mapea
contra alias conocidos. Vacia la tabla y reinserta todo (snapshot completo).
"""
from __future__ import annotations

import csv
import io
import unicodedata
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, insert, select
from sqlalchemy.orm import Session

from ..config import get_settings
from ..models import DimProducto, DimSucursal, Sugerido

settings = get_settings()

# --- Normalizacion de cabeceras -------------------------------------------------

def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    for ch in ("?", "(", ")", "%", "$"):
        s = s.replace(ch, "")
    for ch in (" ", "-", "/", "."):
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


# Mapa: nombre_normalizado_en_archivo -> campo del modelo.
# Se incluyen alias frecuentes del export del BI.
HEADER_ALIASES: dict[str, str] = {
    "producto": "producto",
    "codigo": "producto",
    "codigo_producto": "producto",
    "descripcion": "descripcion",
    "descripcion_corta": "descripcion",
    "sucursal_id": "sucursal_id",
    "sucursalid": "sucursal_id",
    "sucursal": "sucursal_id",
    "id_sucursal": "sucursal_id",
    "nombre_sucursal": "nombre_sucursal",
    "empresa": "empresa",
    "clasificacion_abc": "clasificacion_abc",
    "abc": "clasificacion_abc",
    "clasificacion_abc_agregada": "clasificacion_abc_agregada",
    "abc_agregada": "clasificacion_abc_agregada",
    "clasificacion_abc_agg": "clasificacion_abc_agregada",
    "sucursales_origen_cd": "sucursales_origen_cd",
    "sucursales_origen": "sucursales_origen_cd",
    "proveedor": "proveedor",
    "filtro1_final": "filtro1_final",
    "filtro1": "filtro1_final",
    "marca": "filtro1_final",
    "tipo_origen": "tipo_origen",
    "es_importado": "es_importado",
    "unidad_medida": "unidad_medida",
    "unidad_de_medida": "unidad_medida",
    "lead_time_dias": "lead_time_dias",
    "lt_efectivo": "lt_efectivo",
    "lt_cd_a_sucursal_dias": "lt_cd_a_sucursal_dias",
    "lt_origen": "lt_origen",
    "abastece_cd": "abastece_cd",
    "prioridad_cd": "prioridad_cd",
    "comprar_en_el_cd": "comprar_en_el_cd",
    "tiene_stock_cd": "tiene_stock_cd",
    "demanda_mensual": "demanda_mensual",
    "demanda_diaria": "demanda_diaria",
    "desv_std_mensual": "desv_std_mensual",
    "stock_seguridad": "stock_seguridad",
    "stock_de_seguridad": "stock_seguridad",
    "punto_de_pedido": "punto_de_pedido",
    "costo_unitario": "costo_unitario",
    "pedir": "pedir",
    "reemplazos": "reemplazos",
    "sugerido_suc": "sugerido_suc",
    "stock_activo_suc": "stock_activo_suc",
    "stock_en_transito_suc": "stock_en_transito_suc",
    "stock_en_cd": "stock_en_cd",
    "sugerido_traslado": "sugerido_traslado",
    "sugerido_compra_neto": "sugerido_compra_neto",
    "total_sugerido_suc": "total_sugerido_suc",
    "total_valor_sugerido_clp": "total_valor_sugerido_clp",
    "total_valor_sugerido_suc_clp": "total_valor_sugerido_clp",
    "valor_sugerido_suc_clp": "total_valor_sugerido_clp",
    "valor_sugerido_clp": "total_valor_sugerido_clp",
    "pedir_flag": "pedir_flag",
    # Traslado lateral: texto "N unidades desde X; M desde Y" (medida del BI).
    "trasladar_desde": "trasladar_desde",
    "traslado_desde_otras_sucursales": "trasladar_desde",
    # Stock por bodega/sucursal (columnas fisicas del BI, expandidas por grupo
    # de reemplazo). El comprador las usa para decidir traslado en vez de compra.
    "stock_linderos": "stock_linderos",
    "stock_curico": "stock_curico",
    "stock_talca": "stock_talca",
    "stock_rancagua": "stock_rancagua",
    "stock_diez_de_julio_2": "stock_diez_de_julio_2",
    "stock_chillan": "stock_chillan",
    "stock_cd_repuestos": "stock_cd_repuestos",
    "stock_brasil_18": "stock_brasil_18",
    "stock_placilla": "stock_placilla",
    "stock_chillan_viejo": "stock_chillan_viejo",
    "stock_talca_2": "stock_talca_2",
    # Precios FORD (columnas del BI que cruzan el codigo del sugerido con la tabla Precios).
    "precio_flota_ford": "precio_flota_ford",
    "precio_dealer_ford": "precio_dealer_ford",
    "precio_publico_ford": "precio_publico_ford",
    "precio_publico_iva_ford": "precio_publico_iva_ford",
    "precio_reposicion_ford": "precio_reposicion_ford",
    "precio_urgente_vor_ford": "precio_urgente_vor_ford",
    "precio_promociones_ford": "precio_promociones_ford",
    "precio_urgente_recargo15_ford": "precio_urgente_recargo15_ford",
    # Nombres tal como aparecen en el VISUAL de Power BI (medidas):
    "total_sugerido": "total_sugerido_suc",
    "stock_activo": "stock_activo_suc",
    "stock_en_transito": "stock_en_transito_suc",
    "comprar_en_cd": "comprar_en_el_cd",
}

# Tipos por campo para castear valores del archivo.
INT_FIELDS = {
    "lead_time_dias", "lt_efectivo", "lt_cd_a_sucursal_dias", "prioridad_cd",
    "stock_seguridad", "punto_de_pedido",
    "stock_linderos", "stock_curico", "stock_talca", "stock_rancagua",
    "stock_diez_de_julio_2", "stock_chillan", "stock_cd_repuestos",
    "stock_brasil_18", "stock_placilla", "stock_chillan_viejo", "stock_talca_2",
    "precio_flota_ford", "precio_dealer_ford", "precio_publico_ford",
    "precio_publico_iva_ford", "precio_reposicion_ford", "precio_urgente_vor_ford",
    "precio_promociones_ford", "precio_urgente_recargo15_ford",
}
FLOAT_FIELDS = {
    "demanda_mensual", "demanda_diaria", "desv_std_mensual", "costo_unitario",
    "sugerido_suc", "stock_activo_suc", "stock_en_transito_suc", "stock_en_cd",
    "sugerido_traslado", "sugerido_compra_neto", "total_sugerido_suc",
    "total_valor_sugerido_clp",
}
BOOL_FIELDS = {"es_importado", "tiene_stock_cd"}


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace("%", "")
    # Formato chileno: miles con punto, decimal con coma -> normalizar.
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v: Any) -> int | None:
    f = _to_float(v)
    return int(round(f)) if f is not None else None


def _to_bool(v: Any) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("si", "sí", "true", "verdadero", "1", "x"):
        return True
    if s in ("no", "false", "falso", "0"):
        return False
    return None


def _cast(field: str, value: Any) -> Any:
    if field in INT_FIELDS:
        return _to_int(value)
    if field in FLOAT_FIELDS:
        return _to_float(value)
    if field in BOOL_FIELDS:
        return _to_bool(value)
    # texto
    if value is None:
        return None
    return str(value).strip() or None


# --- Lectura de archivos --------------------------------------------------------

def _rows_from_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows)]
    data = [list(r) for r in rows]
    wb.close()
    return headers, data


def _rows_from_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Dialecto determinista: el extractor de Power BI emite coma + comillas dobles
    # estandar. csv.Sniffer adivinaba el dialecto sobre los primeros 4096 bytes y
    # con campos entrecomillados mixtos (ej. descripciones con 15"/16) a veces
    # descuadraba filas. Solo se detecta el delimitador (un CSV manual chileno
    # puede venir con ';') mirando la linea de cabeceras.
    primera_linea = text.split("\n", 1)[0]
    delim = ";" if primera_linea.count(";") > primera_linea.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim, quotechar='"', doublequote=True)
    all_rows = list(reader)
    if not all_rows:
        return [], []
    return all_rows[0], all_rows[1:]


# --- Carga principal ------------------------------------------------------------

def procesar_registros(
    registros: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    """Dada una lista de registros (dicts {cabecera_cruda: valor}), normaliza las
    cabeceras, las mapea a campos del modelo y castea los valores.

    Es la logica compartida entre la carga por Excel/CSV y la carga desde Power BI.
    Devuelve (filas_mapeadas, columnas_detectadas, columnas_ignoradas).
    """
    mapping: dict[str, str] = {}
    detectadas: list[str] = []
    ignoradas: list[str] = []
    vistos: set[str] = set()
    for reg in registros:
        for h in reg.keys():
            if h in vistos:
                continue
            vistos.add(h)
            field = HEADER_ALIASES.get(_norm(h))
            if field:
                mapping[h] = field
                detectadas.append(f"{h} -> {field}")
            elif h:
                ignoradas.append(h)

    filas: list[dict[str, Any]] = []
    for reg in registros:
        valores: dict[str, Any] = {}
        for h, field in mapping.items():
            if h in reg:
                valores[field] = _cast(field, reg[h])
        filas.append(valores)
    return filas, detectadas, ignoradas


def persistir_filas(
    db: Session,
    filas: list[dict[str, Any]],
    detectadas: list[str],
    ignoradas: list[str],
) -> dict:
    """Reemplaza el snapshot de `sugerido` (y los catalogos) con las filas dadas.

    Cada fila es un dict {campo_modelo: valor_casteado}. Debe incluir al menos
    `producto` y `sucursal_id`.
    """
    campos = set().union(*[f.keys() for f in filas]) if filas else set()
    if "producto" not in campos:
        raise ValueError("No se encontro la columna 'producto' en los datos.")
    if "sucursal_id" not in campos:
        raise ValueError("No se encontro la columna de sucursal en los datos.")

    advertencias: list[str] = []
    if ignoradas:
        advertencias.append("Columnas ignoradas (sin mapeo): " + ", ".join(ignoradas[:20]))

    tenant = settings.default_tenant_id

    registros_sugerido: list[dict[str, Any]] = []
    productos_vistos: dict[str, dict] = {}
    sucursales_vistas: dict[str, dict] = {}
    saltadas = 0

    for valores in filas:
        if not valores.get("producto") or not valores.get("sucursal_id"):
            saltadas += 1
            continue
        registros_sugerido.append({**valores, "tenant_id": tenant})

        p = valores["producto"]
        if p not in productos_vistos:
            productos_vistos[p] = {
                "producto": p,
                "tenant_id": tenant,
                "descripcion": valores.get("descripcion"),
                "filtro1_final": valores.get("filtro1_final"),
                "unidad_medida": valores.get("unidad_medida"),
                "costo_unitario": valores.get("costo_unitario"),
                "proveedor": valores.get("proveedor"),
                "es_importado": valores.get("es_importado"),
            }
        s = valores["sucursal_id"]
        if s not in sucursales_vistas:
            sucursales_vistas[s] = {
                "sucursal_id": s,
                "tenant_id": tenant,
                "nombre": valores.get("nombre_sucursal"),
                "abastece_desde_cd": valores.get("abastece_cd"),
                "prioridad_cd": valores.get("prioridad_cd"),
            }

    # Guardrail pre-reemplazo: nunca pisar un snapshot sano con uno sospechosamente
    # chico (extraccion cortada a la mitad, CSV corrupto, modelo equivocado abierto).
    previas = (
        db.execute(
            select(func.count()).select_from(Sugerido).where(Sugerido.tenant_id == tenant)
        ).scalar()
        or 0
    )
    minimo = int(previas * settings.sync_min_ratio_filas)
    if previas and len(registros_sugerido) < minimo:
        raise ValueError(
            f"Carga abortada: llegaron {len(registros_sugerido)} filas validas y el snapshot "
            f"actual tiene {previas} (minimo aceptado: {minimo}). Se conserva el snapshot "
            "anterior. Si la baja es real, ajustar SYNC_MIN_RATIO_FILAS."
        )

    # Inserts en multi-fila (un INSERT con muchas VALUES por lote). Con pg8000 esto es
    # MUCHO mas rapido que executemany (que iria fila por fila por la red). chunk=500
    # mantiene los parametros por debajo del limite de Postgres (~65k).
    def _bulk(model, registros: list[dict], chunk: int = 500) -> None:
        for i in range(0, len(registros), chunk):
            lote = registros[i : i + chunk]
            if lote:
                db.execute(insert(model).values(lote))

    # Reemplazo total (snapshot) en una sola transaccion: si cualquier insert falla
    # (corte de red, timeout del pooler), el rollback conserva el snapshot anterior.
    try:
        db.execute(delete(Sugerido).where(Sugerido.tenant_id == tenant))
        _bulk(Sugerido, registros_sugerido)
        db.execute(delete(DimProducto).where(DimProducto.tenant_id == tenant))
        db.execute(delete(DimSucursal).where(DimSucursal.tenant_id == tenant))
        _bulk(DimProducto, list(productos_vistos.values()))
        _bulk(DimSucursal, list(sucursales_vistas.values()))
        db.commit()
    except Exception:
        db.rollback()
        raise

    if saltadas:
        advertencias.append(f"{saltadas} fila(s) sin producto/sucursal fueron omitidas.")

    # Chequeos de calidad post-carga (avisan, no bloquean).
    umbral = settings.sync_umbral_sugerido_unidades
    anomalos = sorted(
        {
            str(r["producto"])
            for r in registros_sugerido
            if (r.get("total_sugerido_suc") or 0) > umbral
        }
    )
    if anomalos:
        advertencias.append(
            f"{len(anomalos)} producto(s) con Total Sugerido > {umbral:,} unidades "
            "(posible unidad de medida corrupta, ej. mL): " + ", ".join(anomalos[:10])
        )
    con_sugerido = [r for r in registros_sugerido if (r.get("total_sugerido_suc") or 0) > 0]
    sin_prov = sum(1 for r in con_sugerido if not r.get("proveedor"))
    if con_sugerido and sin_prov:
        pct = round(100 * sin_prov / len(con_sugerido))
        advertencias.append(
            f"{sin_prov} de {len(con_sugerido)} filas con sugerido ({pct}%) no tienen "
            "proveedor asignado (caeran al carro 'Sin proveedor asignado')."
        )

    return {
        "filas_cargadas": len(registros_sugerido),
        "productos": len(productos_vistos),
        "sucursales": len(sucursales_vistas),
        "columnas_detectadas": detectadas,
        "advertencias": advertencias,
    }


def cargar_sugerido(db: Session, filename: str, content: bytes) -> dict:
    """Parsea el archivo Excel/CSV y reemplaza el contenido de la tabla `sugerido`."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        headers, data = _rows_from_csv(content)
    elif name.endswith((".xlsx", ".xlsm")):
        headers, data = _rows_from_xlsx(content)
    else:
        raise ValueError("Formato no soportado. Usa .xlsx o .csv")

    registros = [dict(zip(headers, raw)) for raw in data]
    filas, detectadas, ignoradas = procesar_registros(registros)
    resumen = persistir_filas(db, filas, detectadas, ignoradas)

    # Historia + alertas. Va DESPUES del commit de la carga y no propaga errores:
    # un problema guardando la foto del dia no puede dejar la plataforma sin datos.
    from . import snapshot_service

    resumen["post_carga"] = snapshot_service.post_carga(db)
    return resumen
